"""
Support for Microsoft Information Protection (MIP) labels in Excel files with openpyxl

"""

from datetime import datetime
from enum import Enum
import json
from glob import glob
import os
import logging
from traceback import extract_stack
from typing import Optional, Dict, Union

from pydantic import BaseModel, Field, ConfigDict
from openpyxl import Workbook, load_workbook
from openpyxl.packaging.custom import StringProperty

from json_toolbox import DateTimeEncoder

logger = logging.getLogger()

DEFAULT_SENSITIVITY_LABELS_DEFINITION = (
    "sensitivity_model/sensitivity_labels_definition_with_openpyxl.json"
)
DEFAULT_SENSITIVITY_TEMPLATES = "sensitivity_model"


class MsoAssignmentMethod(Enum):
    """Enum representing the assignment method of a sensitivity label."""

    Not_Set = -1
    Standard = 0
    Privileged = 1
    Auto = 2


class MSIP_Label(BaseModel):
    """Represents a Microsoft Information Protection (MIP) label.

    This class models the structure of sensitivity labels as defined in the MIP framework, facilitating the handling
    of such labels in Excel documents through the openpyxl library.

    Attributes:
        LabelId (str): Unique identifier for the label.
        LabelName (str): Human-readable name of the label.
        ActionId (Optional[str]): Identifier of the action associated with the label.
        AssignmentMethod (Optional[Union[str, int]]): Method used to assign the label.
        ContentBits (Optional[int]): Bit flags representing content attributes.
        IsEnabled (Optional[bool]): Flag indicating whether the label is enabled.
        Justification (Optional[str]): Justification for label assignment, if any.
        SetDate (Optional[datetime]): Timestamp of when the label was set.
        SiteId (Optional[str]): Identifier for the site associated with the label.
    """

    model_config = ConfigDict(populate_by_name=True)
    LabelId: str
    LabelName: str = Field(..., alias="Name")
    ActionId: Optional[str]
    AssignmentMethod: Optional[Union[str, int]] = Field(alias="Method")
    ContentBits: Optional[int]
    IsEnabled: Optional[bool] = Field(alias="Enabled")
    Justification: Optional[str] = Field(None)
    SetDate: Optional[datetime]
    SiteId: Optional[str]


class MSIP_Manager:
    """Manages Microsoft Information Protection (MIP) labels within an Excel workbook.

    Provides functionality to retrieve and set MIP labels in the workbook's custom document properties.

    Attributes:
        workbook (Workbook): The openpyxl Workbook instance.
    """

    def __init__(self, workbook: Workbook):
        self.workbook = workbook

    def getlabel(self) -> Optional[MSIP_Label]:
        """Retrieves the MIP label from the workbook's custom document properties.

        Returns:
            An instance of MSIP_Label if a label is found, otherwise None.
        """
        custom_doc_props = getattr(self.workbook, "custom_doc_props", None)
        if not custom_doc_props:
            return None
        msip_info = dict()
        for prop in custom_doc_props.props:
            logger.debug(
                f"GetProperty :  {type(prop) =}{prop.name = }: {prop.value = }"
            )
            if prop.name.startswith("MSIP_Label"):
                prop_name_parts = prop.name.split("_")
                label_id = prop_name_parts[-2]
                attr = prop_name_parts[-1]
                msip_info["LabelId"] = label_id
                msip_info[attr] = prop.value
        msip_label = MSIP_Label.model_validate(msip_info)
        return msip_label

    def setlabel(self, msip_label: MSIP_Label, justification: Optional[str] = None):
        """Sets the MIP label to the workbook's custom document properties.

        Args:
            msip_label (MSIP_Label): The MIP label to be applied to the workbook.
            justification (Optional[str]): Justification for applying the label, if any.
        """
        custom_doc_props = getattr(self.workbook, "custom_doc_props")
        # change to now
        if justification:
            msip_label.Justification = justification
        msip_label.SetDate = datetime.now()
        for prop_name, prop_value in msip_label.model_dump(
            by_alias=True, exclude={"LabelId"}
        ).items():
            prop = StringProperty(
                name=f"MSIP_Label_{msip_label.LabelId}_{prop_name}",
                value=str(prop_value),
            )
            logger.debug(f"SetProperty : {type(prop) =}{prop.name = }: {prop.value = }")
            custom_doc_props.append(prop)


class MSIP_Configuration:
    """Manages the configuration of Microsoft Information Protection (MIP) labels.

    Loads, saves, and manipulates sensitivity labels defined in a JSON configuration file.

    Attributes:
        sensitivity_configuration_file (str): Path to the JSON file containing sensitivity label definitions.
        sensitivity_labels (Dict[str, MSIP_Label]): Dictionary mapping label names to MSIP_Label instances.
    """

    def __init__(
        self,
        sensitivity_configuration_file: str = DEFAULT_SENSITIVITY_LABELS_DEFINITION,
    ):
        self.sensitivity_configuration_file = sensitivity_configuration_file
        self.sensitivity_labels: Dict[str, MSIP_Label] = {}

    def load(self):
        """Loads sensitivity label definitions from the configuration file."""
        with open(self.sensitivity_configuration_file, "r") as fh_in:
            for label_name, label_info in json.load(fh_in).items():
                msip_label = MSIP_Label.model_validate(label_info, from_attributes=True)
                self.add_sensitivity_label(label_name, msip_label)
        return self

    def save(self):
        """Saves sensitivity label definitions to the configuration file."""
        dump_json = dict()
        for label_name, msip_label in self.sensitivity_labels.items():
            dump_json[label_name] = msip_label.model_dump()

        with open(self.sensitivity_configuration_file, "w") as fh_out:
            json.dump(dump_json, fh_out, indent=4, cls=DateTimeEncoder)

    def add_sensitivity_label(self, label_name: str, msip_label: MSIP_Label):
        """Adds a sensitivity label to the configuration."""
        self.sensitivity_labels[label_name] = msip_label

    def get_sensitivity_label(self, label_name: str) -> MSIP_Label:
        """Retrieves a sensitivity label by name."""
        return self.sensitivity_labels[label_name]

    def labels(self) -> Dict[str, MSIP_Label].keys:  # type: ignore
        """Returns the names of all configured sensitivity labels."""
        return self.sensitivity_labels.keys()


def create_sensitivity_label_definition(
    extract_from: str = DEFAULT_SENSITIVITY_TEMPLATES,
    sensitivity_configuration_file: str = DEFAULT_SENSITIVITY_LABELS_DEFINITION,
) -> MSIP_Configuration:
    """
    Creates and saves a configuration of sensitivity labels based on label information extracted from Excel files.

    This function scans a directory for Excel files, extracts MIP label information from each file, and compiles
    a comprehensive configuration of all labels found. This configuration is then saved to a JSON file for future use.

    Args:
        extract_from (str): The directory path from which to extract sensitivity labels from Excel files. Defaults to
                            the value of DEFAULT_SENSITIVITY_TEMPLATES.
        sensitivity_configuration_file (str): The file path to save the extracted sensitivity label configuration. Defaults
                                              to the value of DEFAULT_SENSITIVITY_LABELS_DEFINITION.

    Returns:
        MSIP_Configuration: An instance of MSIP_Configuration loaded with the compiled sensitivity labels.
    """
    msip_configuration = MSIP_Configuration(sensitivity_configuration_file)
    for filename in glob(os.path.join(extract_from, "*.xlsx")):
        label_name, _ = os.path.splitext(os.path.basename(filename))
        msip_label = get_label_from_file(filename)
        if msip_label:
            msip_configuration.add_sensitivity_label(label_name, msip_label)
    msip_configuration.save()
    return msip_configuration


def get_label_from_file(filename: str) -> Optional[MSIP_Label]:
    """
    Extracts the sensitivity label information from a given Excel file.

    This function opens an Excel workbook and retrieves any sensitivity label information stored within its
    custom document properties. It is useful for reading label data from individual files.

    Args:
        filename (str): The path to the Excel file from which to extract the sensitivity label.

    Returns:
        Optional[MSIP_Label]: An instance of MSIP_Label containing the extracted label information, if found. Returns
                              None if no label information is present in the file.
    """
    wb = load_workbook(filename)
    msip_manager = MSIP_Manager(wb)
    return msip_manager.getlabel()


def set_label_to_workbook(wb: Workbook, label: MSIP_Label):
    """
    Applies a sensitivity label to an openpyxl Workbook object.

    This function sets the provided MIP label information as custom document properties of the given Workbook.
    It's utilized to programmatically apply sensitivity labels to Excel files being managed through openpyxl.

    Args:
        wb (Workbook): The openpyxl Workbook instance to which the label will be applied.
        label (MSIP_Label): The sensitivity label to apply to the workbook.

    No return value.
    """
    msip_manager = MSIP_Manager(wb)
    msip_manager.setlabel(label)


def set_label_to_file(filename: str, label: MSIP_Label):
    """
    Applies a sensitivity label to an Excel file.

    Opens the specified Excel file, applies the given sensitivity label to it, and saves the changes. This is a
    convenience function that combines file loading, label application, and saving operations into one step.

    Args:
        filename (str): The path to the Excel file to which the label will be applied.
        label (MSIP_Label): The sensitivity label to apply to the file.

    No return value.
    """
    wb = load_workbook(filename)
    set_label_to_workbook(wb, label)
    wb.save(filename)
